from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta

from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError as DjangoValidationError

from userauths.models import Patient, User, Role, Hospital, Department, PatientVisit
from userauths.serializer import PatientSerializer, PatientReferralSerializer, PatientCreateSerializer, PatientUpdateSerializer
from userauths.permissions import PatientAccessControl


class PatientViewSet(viewsets.ModelViewSet):
    """
    Patient management ViewSet.
    - Receptionists: CRUD for patients at their assigned hospital
    - Doctors/Nurses: Read-only access to patients at their hospital
    - Admins: Full access to all patients
    """
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return PatientCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return PatientUpdateSerializer
        return PatientSerializer

    def get_queryset(self):
        user = self.request.user
        role = user.role.name if user.role else None
        action = self.action
        
        # Get search query first
        search = self.request.query_params.get('search', None)
        # Base access control
        # Admins and ministry see all patients by default
        if role in ['admin', 'ministry_admin']:
            queryset = Patient.objects.all()
        # For retrieve/detail view, return all patients so access control
        # logic can properly evaluate cross-hospital access (403, not 404)
        elif action == 'retrieve':
            queryset = Patient.objects.all()
        # If there's a search, only show:
        #   - Patients registered at this hospital
        #   - Patients with an actual REFERRAL visit at this hospital
        #   - Patients with an actual REFERRAL appointment at this hospital
        elif search and role in ['receptionist', 'doctor', 'nurse', 'hospital_admin', 'triage']:
            queryset = Patient.objects.filter(
                Q(hospital=user.hospital) |
                Q(visits__hospital=user.hospital, visits__visit_type='referral') |
                Q(appointments__hospital=user.hospital, appointments__is_referral=True)
            ).distinct()
        # Hospital-level staff see only their hospital's patients for regular browsing
        elif user.hospital:
            queryset = Patient.objects.filter(hospital=user.hospital)
        # District admins see patients across their district's hospitals
        elif role == 'district_admin' and user.district:
            queryset = Patient.objects.filter(hospital__district=user.district)
        else:
            queryset = Patient.objects.none()

        # Apply Search filtering
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(other_names__icontains=search) |
                Q(patient_id__icontains=search) |
                Q(phone__icontains=search) |
                Q(national_id__icontains=search) |
                Q(insurance_number__icontains=search) |
                Q(email__icontains=search)
            )

        # Filter by gender
        gender = self.request.query_params.get('gender', None)
        if gender:
            queryset = queryset.filter(gender=gender)

        # Filter by blood type
        blood_type = self.request.query_params.get('blood_type', None)
        if blood_type:
            queryset = queryset.filter(blood_type=blood_type)

        # Filter by status
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by insurance provider
        insurance = self.request.query_params.get('insurance', None)
        if insurance:
            queryset = queryset.filter(insurance_provider__icontains=insurance)

        # Filter incomplete profiles (missing national_id — nurse quick-registered)
        incomplete = self.request.query_params.get('incomplete', None)
        if incomplete == 'true':
            queryset = queryset.filter(
                Q(national_id__isnull=True) | Q(national_id='')
            )

        return queryset.select_related('hospital', 'registered_by')

    def retrieve(self, request, *args, **kwargs):
        """Retrieve a single patient with access control and audit logging."""
        instance = self.get_object()
        user = request.user
        role = user.role.name if user.role else None

        # Admins bypass access control (already handled by get_object)
        if role in ('admin', 'ministry_admin'):
            PatientAccessControl._log(user, instance, 'view', 'admin', 'allowed', request)
            serializer = self.get_serializer(instance)
            return Response(serializer.data)

        # Check access control for non-admins
        result = PatientAccessControl.can_access_patient(
            user, instance, action='view', request=request
        )

        if result['allowed']:
            # Use limited serializer for cross-hospital referral access
            if result['access_type'] == 'cross_hospital':
                serializer = PatientReferralSerializer(instance)
            else:
                serializer = self.get_serializer(instance)
            return Response(serializer.data)
        else:
            return Response(
                {'error': result['message']},
                status=status.HTTP_403_FORBIDDEN
            )

    def get_object(self):
        """Default object lookup — used by update/delete."""
        return super().get_object()

    @action(detail=True, methods=['post'], url_path='emergency_access')
    def emergency_access(self, request, pk=None):
        """
        Break-the-glass emergency access to a patient record.
        POST body: { "justification": "Patient in cardiac arrest..." }
        """
        user = request.user
        justification = request.data.get('justification', '').strip()

        if not justification or len(justification) < 10:
            return Response(
                {'error': 'A detailed justification (at least 10 characters) is required for emergency access.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        patient = self.get_object()
        result = PatientAccessControl.can_access_patient(
            user, patient, action='emergency', request=request, justification=justification
        )

        if result['allowed']:
            serializer = self.get_serializer(patient)
            return Response({
                'allowed': True,
                'access_type': result['access_type'],
                'message': 'Emergency access granted. This action has been logged and will be reviewed.',
                'patient': serializer.data,
            })
        else:
            return Response(
                {'error': result['message']},
                status=status.HTTP_403_FORBIDDEN
            )

    @action(detail=True, methods=['post'], url_path='refer')
    def refer_patient(self, request, pk=None):
        """
        Send a patient referral to another hospital.

        POST body:
        {
            "referred_to_hospital": <hospital_id>,
            "department": <department_id> (optional),
            "doctor": <user_id> (optional),
            "reason": "<chief complaint / referral reason>",
            "notes": "<additional notes>" (optional)
        }

        Creates a PatientVisit(visit_type='referral') at the current hospital
        with referred_to_hospital set, which grants that hospital's care-team
        access to the patient record under the cross_hospital access rule.
        """
        actor = request.user
        role  = actor.role.name if actor.role else None

        allowed_roles = ('doctor', 'receptionist', 'hospital_admin', 'admin', 'ministry_admin')
        if role not in allowed_roles:
            return Response(
                {'error': 'Only doctors, receptionists and hospital admins can send referrals.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        if not actor.hospital:
            return Response(
                {'error': 'You are not assigned to a hospital.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        patient = self.get_object()

        # Validate required fields
        referred_to_hospital_id = request.data.get('referred_to_hospital')
        reason = request.data.get('reason', '').strip()

        if not referred_to_hospital_id:
            return Response({'error': 'Please select the hospital you are referring this patient to.'}, status=status.HTTP_400_BAD_REQUEST)
        if not reason:
            return Response({'error': 'A referral reason is required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Prevent referring to the same hospital
        try:
            dest_hospital = Hospital.objects.get(id=referred_to_hospital_id)
        except Hospital.DoesNotExist:
            return Response({'error': 'Destination hospital not found.'}, status=status.HTTP_404_NOT_FOUND)

        if dest_hospital == actor.hospital:
            return Response({'error': 'You cannot refer a patient to your own hospital.'}, status=status.HTTP_400_BAD_REQUEST)

        # Optional fields
        department_id = request.data.get('department')
        doctor_id     = request.data.get('doctor')
        notes         = request.data.get('notes', '')

        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                pass

        referred_doctor_name = ''
        if doctor_id:
            try:
                dest_doctor = User.objects.get(id=doctor_id, hospital=dest_hospital)
                referred_doctor_name = dest_doctor.full_name or dest_doctor.email
            except User.DoesNotExist:
                pass

        # Create the referral visit record at the current hospital
        visit = PatientVisit.objects.create(
            patient=patient,
            hospital=actor.hospital,
            department=department,
            doctor=actor if role == 'doctor' else None,
            registered_by=actor,
            visit_type='referral',
            chief_complaint=reason,
            visit_date=timezone.now(),
            status='referred_out',
            referred_to_hospital=dest_hospital,
            referred_to_doctor=referred_doctor_name,
            discharge_notes=notes,
        )

        # Audit log
        PatientAccessControl._log(actor, patient, 'edit', 'same_hospital', 'allowed', request)

        return Response({
            'message': f'{patient.full_name} has been referred to {dest_hospital.name}.',
            'visit_id': visit.id,
            'referred_to': dest_hospital.name,
            'referred_to_id': dest_hospital.id,
        }, status=status.HTTP_201_CREATED)

    def create(self, request, *args, **kwargs):
        user = request.user
        role = user.role.name if user.role else None

        # Receptionists, nurses, hospital_admin, and admin can register patients
        allowed_roles = ['receptionist', 'nurse', 'triage', 'hospital_admin', 'admin', 'ministry_admin']
        if role not in allowed_roles:
            return Response(
                {'error': 'You do not have permission to register patients.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Receptionist / nurse / triage must be assigned to a hospital
        if role in ['receptionist', 'nurse', 'triage'] and not user.hospital:
            return Response(
                {'error': 'You are not assigned to any hospital. Contact your administrator.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Auto-assign hospital and registered_by
        hospital = user.hospital
        if role in ['admin', 'ministry_admin']:
            # Admin can optionally specify a hospital
            hospital_id = request.data.get('hospital')
            if hospital_id:
                from userauths.models import Hospital
                try:
                    hospital = Hospital.objects.get(id=hospital_id)
                except Hospital.DoesNotExist:
                    return Response({'error': 'Invalid hospital ID.'}, status=status.HTTP_400_BAD_REQUEST)
            elif not hospital:
                return Response(
                    {'error': 'Please specify a hospital for this patient.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        patient = serializer.save(hospital=hospital, registered_by=user)

        return Response({
            'message': f'Patient {patient.full_name} registered successfully.',
            'patient': PatientSerializer(patient).data
        }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        user = request.user
        role = user.role.name if user.role else None
        allowed_roles = ['receptionist', 'hospital_admin', 'admin', 'ministry_admin']
        if role not in allowed_roles:
            return Response(
                {'error': 'You do not have permission to edit patients.'},
                status=status.HTTP_403_FORBIDDEN
            )

        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response({
            'message': 'Patient updated successfully.',
            'patient': PatientSerializer(instance).data
        })

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Return patient statistics for the dashboard."""
        queryset = self.get_queryset()
        now = timezone.now()
        today = now.date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)

        total = queryset.count()
        active = queryset.filter(status='active').count()
        inactive = queryset.filter(status='inactive').count()
        today_count = queryset.filter(created_at__date=today).count()
        week_count = queryset.filter(created_at__date__gte=week_ago).count()
        month_count = queryset.filter(created_at__date__gte=month_ago).count()

        gender_dist = list(
            queryset.values('gender').annotate(count=Count('id')).order_by('-count')
        )
        blood_dist = list(
            queryset.exclude(blood_type='unknown').values('blood_type').annotate(count=Count('id')).order_by('-count')
        )
        insurance_dist = list(
            queryset.exclude(insurance_provider__isnull=True).exclude(insurance_provider='')
            .values('insurance_provider').annotate(count=Count('id')).order_by('-count')
        )

        # Monthly registration trend (last 6 months)
        monthly_trend = []
        for i in range(5, -1, -1):
            month_start = (now - timedelta(days=i * 30)).replace(day=1).date()
            if i > 0:
                next_month = (now - timedelta(days=(i - 1) * 30)).replace(day=1).date()
            else:
                next_month = today + timedelta(days=1)
            count = queryset.filter(created_at__date__gte=month_start, created_at__date__lt=next_month).count()
            monthly_trend.append({
                'month': month_start.strftime('%b %Y'),
                'count': count
            })

        # Incomplete profiles: patients registered without a national ID (quick-registered by nurse)
        incomplete_count = queryset.filter(
            Q(national_id__isnull=True) | Q(national_id='')
        ).count()

        return Response({
            'total_patients': total,
            'active_patients': active,
            'inactive_patients': inactive,
            'registered_today': today_count,
            'registered_this_week': week_count,
            'registered_this_month': month_count,
            'incomplete_profiles': incomplete_count,
            'gender_distribution': gender_dist,
            'blood_type_distribution': blood_dist,
            'insurance_distribution': insurance_dist,
            'monthly_trend': monthly_trend,
        })

    @action(detail=True, methods=['post'], url_path='create_portal_account')
    def create_portal_account(self, request, pk=None):
        """
        Receptionist creates a patient portal (login) account for a registered patient.
        Required body: { "password": "...", "password2": "..." }
        """
        actor = request.user
        allowed_roles = ('admin', 'receptionist', 'hospital_admin')
        if not actor.role or actor.role.name not in allowed_roles:
            return Response({'error': 'Only receptionists can create patient portal accounts.'}, status=status.HTTP_403_FORBIDDEN)

        patient = self.get_object()

        if patient.user:
            return Response({'error': 'This patient already has a portal account.'}, status=status.HTTP_400_BAD_REQUEST)

        if not patient.email:
            return Response({'error': 'Patient must have an email address to create a portal account.'}, status=status.HTTP_400_BAD_REQUEST)

        password  = request.data.get('password', '')
        password2 = request.data.get('password2', '')

        if not password or not password2:
            return Response({'error': 'Password and confirmation are required.'}, status=status.HTTP_400_BAD_REQUEST)

        if password != password2:
            return Response({'error': 'Passwords do not match.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            validate_password(password)
        except DjangoValidationError as e:
            return Response({'error': ' '.join(e.messages)}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(email=patient.email).exists():
            return Response({'error': f'A user account with the email {patient.email} already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        patient_role, _ = Role.objects.get_or_create(name='patient')

        user = User.objects.create(
            email=patient.email,
            full_name=patient.full_name,
            phone=patient.phone,
            role=patient_role,
            hospital=patient.hospital,
        )
        user.set_password(password)
        user.save()

        patient.user = user
        patient.save(update_fields=['user'])

        return Response({
            'message': f'Portal account created for {patient.full_name}.',
            'email': patient.email,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='download_template')
    def download_template(self, request):
        """Return a blank Excel template for bulk patient upload."""
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Patients'

        headers = [
            'first_name*', 'last_name*', 'other_names', 'date_of_birth* (YYYY-MM-DD)',
            'gender* (male/female/other)', 'nationality', 'national_id',
            'phone*', 'alt_phone', 'email',
            'address', 'city', 'blood_type (A+/A-/B+/B-/AB+/AB-/O+/O-/unknown)',
            'allergies', 'chronic_conditions', 'disabilities',
            'insurance_provider', 'insurance_number',
            'emergency_contact_name', 'emergency_contact_phone', 'emergency_contact_relationship',
        ]

        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 18)

        # Add one example row
        ws.append([
            'Aminata', 'Kamara', '', '1990-05-15',
            'female', 'Sierra Leonean', 'SL-NIN-12345',
            '+23276000000', '', 'aminata@example.com',
            '12 Main Street', 'Freetown', 'O+',
            'Peanuts', 'Hypertension', '',
            'SLeSHI', 'INS-001',
            'Fatmata Kamara', '+23277000000', 'Mother',
        ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="patient_upload_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='bulk_upload')
    def bulk_upload(self, request):
        """
        Upload an Excel file to register multiple patients at once.
        Skips invalid rows silently. Returns counts of imported and skipped rows.
        """
        import openpyxl
        from datetime import date

        user = request.user
        role = user.role.name if user.role else None
        allowed_roles = ['receptionist', 'nurse', 'triage', 'hospital_admin', 'admin', 'ministry_admin']
        if role not in allowed_roles:
            return Response({'error': 'You do not have permission to upload patients.'}, status=status.HTTP_403_FORBIDDEN)

        hospital = user.hospital
        if role in ['admin', 'ministry_admin']:
            hospital_id = request.data.get('hospital')
            if hospital_id:
                try:
                    hospital = Hospital.objects.get(id=hospital_id)
                except Hospital.DoesNotExist:
                    return Response({'error': 'Invalid hospital ID.'}, status=status.HTTP_400_BAD_REQUEST)
        if not hospital:
            return Response({'error': 'No hospital assigned. Cannot import patients.'}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception:
            return Response({'error': 'Invalid Excel file. Please use the provided template.'}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        GENDER_VALID = {'male', 'female', 'other'}
        BLOOD_VALID  = {'A+','A-','B+','B-','AB+','AB-','O+','O-','unknown'}

        imported = 0
        skipped  = 0

        for row in rows:
            if not any(row):
                continue
            try:
                (first_name, last_name, other_names, dob_raw,
                 gender, nationality, national_id,
                 phone, alt_phone, email,
                 address, city, blood_type,
                 allergies, chronic_conditions, disabilities,
                 insurance_provider, insurance_number,
                 ec_name, ec_phone, ec_relationship) = (list(row) + [None]*21)[:21]

                # Required fields
                if not first_name or not last_name or not phone:
                    skipped += 1
                    continue

                # Parse date
                if isinstance(dob_raw, date):
                    dob = dob_raw
                elif dob_raw:
                    from datetime import datetime
                    try:
                        dob = datetime.strptime(str(dob_raw).strip(), '%Y-%m-%d').date()
                    except ValueError:
                        skipped += 1
                        continue
                else:
                    skipped += 1
                    continue

                gender = (str(gender).strip().lower() if gender else 'other')
                if gender not in GENDER_VALID:
                    gender = 'other'

                blood_type = (str(blood_type).strip() if blood_type else 'unknown')
                if blood_type not in BLOOD_VALID:
                    blood_type = 'unknown'

                Patient.objects.create(
                    hospital=hospital,
                    registered_by=user,
                    first_name=str(first_name).strip(),
                    last_name=str(last_name).strip(),
                    other_names=str(other_names).strip() if other_names else '',
                    date_of_birth=dob,
                    gender=gender,
                    nationality=str(nationality).strip() if nationality else 'Sierra Leonean',
                    national_id=str(national_id).strip() if national_id else None,
                    phone=str(phone).strip(),
                    alt_phone=str(alt_phone).strip() if alt_phone else None,
                    email=str(email).strip() if email else None,
                    address=str(address).strip() if address else None,
                    city=str(city).strip() if city else None,
                    blood_type=blood_type,
                    allergies=str(allergies).strip() if allergies else None,
                    chronic_conditions=str(chronic_conditions).strip() if chronic_conditions else None,
                    disabilities=str(disabilities).strip() if disabilities else None,
                    insurance_provider=str(insurance_provider).strip() if insurance_provider else None,
                    insurance_number=str(insurance_number).strip() if insurance_number else None,
                    emergency_contact_name=str(ec_name).strip() if ec_name else None,
                    emergency_contact_phone=str(ec_phone).strip() if ec_phone else None,
                    emergency_contact_relationship=str(ec_relationship).strip() if ec_relationship else None,
                )
                imported += 1
            except Exception:
                skipped += 1
                continue

        return Response({
            'message': f'Import complete. {imported} patient(s) imported, {skipped} row(s) skipped.',
            'imported': imported,
            'skipped': skipped,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['delete'], url_path='revoke_portal_account')
    def revoke_portal_account(self, request, pk=None):
        """Receptionist removes a patient's portal access."""
        actor = request.user
        allowed_roles = ('admin', 'receptionist', 'hospital_admin')
        if not actor.role or actor.role.name not in allowed_roles:
            return Response({'error': 'Not authorized.'}, status=status.HTTP_403_FORBIDDEN)

        patient = self.get_object()
        if not patient.user:
            return Response({'error': 'This patient has no portal account.'}, status=status.HTTP_400_BAD_REQUEST)

        old_user = patient.user
        patient.user = None
        patient.save(update_fields=['user'])
        old_user.delete()

        return Response({'message': 'Portal account revoked.'}, status=status.HTTP_200_OK)
